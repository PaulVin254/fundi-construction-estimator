# =============================================================================
# FILE: utils/excel_boq_generator.py
# PURPOSE:
#   Generates a professional, multi-tab Excel (.xlsx) Bill of Quantities with
#   active formulas (=D*E), trade formatting, shopping list, and layman guide.
# =============================================================================

import os
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_boq(boq_data: Dict[str, Any], output_filepath: str) -> str:
    """
    Creates a styled, multi-tab Microsoft Excel (.xlsx) document for the BOQ.
    
    Tabs:
      1. Executive Summary
      2. Itemized BOQ (with dynamic Excel formulas)
      3. Shopping List (Material totals)
      4. Layperson Building Guide
    """
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
    trade_header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Soft Blue
    zebra_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # Light Gray
    accent_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")  # Yellow Highlight
    
    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="4B5563")
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    meta = boq_data.get("project_metadata", {})
    trades = boq_data.get("trades", [])

    # -------------------------------------------------------------------------
    # TAB 1: EXECUTIVE SUMMARY
    # -------------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary.cell(row=1, column=1, value="FUNDI CONSTRUCTION ESTIMATOR").font = title_font
    ws_summary.cell(row=2, column=1, value="Residential Construction Cost Estimate & BOQ Summary").font = subtitle_font

    meta_rows = [
        ("House Type:", meta.get("house_type", "").replace("_", " ").title()),
        ("Location / Region:", meta.get("location_name", "Nairobi")),
        ("Plinth Area:", f"{meta.get('size_sqm', 0)} sq. meters"),
        ("Finish Quality Tier:", meta.get("finish_level", "").title()),
        ("Estimate Date:", meta.get("generated_date", "2026-07-23"))
    ]
    
    r = 4
    for label, val in meta_rows:
        ws_summary.cell(row=r, column=1, value=label).font = bold_font
        ws_summary.cell(row=r, column=2, value=val).font = regular_font
        r += 1

    r += 1
    ws_summary.cell(row=r, column=1, value="TRADE COST SUMMARY").font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    r += 1

    headers_s = ["Trade Code", "Trade Name / Category", "Subtotal (KES)"]
    for c_idx, h in enumerate(headers_s, 1):
        cell = ws_summary.cell(row=r, column=c_idx, value=h)
        cell.font = white_bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx==1 else "left")

    summary_start_row = r + 1
    r += 1
    for t in trades:
        ws_summary.cell(row=r, column=1, value=t.get("trade_code")).font = bold_font
        ws_summary.cell(row=r, column=2, value=f"{t.get('icon', '')} {t.get('trade_name')}").font = regular_font
        amt_cell = ws_summary.cell(row=r, column=3, value=t.get("trade_total"))
        amt_cell.font = bold_font
        amt_cell.number_format = '#,##0'
        r += 1

    summary_end_row = r - 1

    # Contingency & Grand Total
    ws_summary.cell(row=r, column=2, value="Subtotal Materials & Labor:").font = bold_font
    sub_cell = ws_summary.cell(row=r, column=3, value=f"=SUM(C{summary_start_row}:C{summary_end_row})")
    sub_cell.font = bold_font
    sub_cell.number_format = '#,##0'
    r += 1

    ws_summary.cell(row=r, column=2, value="Contingency & Buffer (10%):").font = bold_font
    cont_cell = ws_summary.cell(row=r, column=3, value=f"=C{r-1}*0.10")
    cont_cell.font = bold_font
    cont_cell.number_format = '#,##0'
    r += 1

    ws_summary.cell(row=r, column=2, value="ESTIMATED GRAND TOTAL:").font = Font(name="Calibri", size=12, bold=True)
    gt_cell = ws_summary.cell(row=r, column=3, value=f"=C{r-2}+C{r-1}")
    gt_cell.font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
    gt_cell.fill = accent_fill
    gt_cell.number_format = '#,##0'

    # -------------------------------------------------------------------------
    # TAB 2: ITEMIZED BOQ (WITH FORMULAS)
    # -------------------------------------------------------------------------
    ws_boq = wb.create_sheet(title="Itemized BOQ")
    ws_boq.views.sheetView[0].showGridLines = True

    boq_headers = ["Item Code", "Material / Work Description", "Unit", "Quantity", "Unit Rate (KES)", "Total Amount (KES)", "Layperson Explanation & Purpose"]
    for c_idx, h in enumerate(boq_headers, 1):
        cell = ws_boq.cell(row=1, column=c_idx, value=h)
        cell.font = white_bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx in [1,3,4,5,6] else "left")

    row_idx = 2
    for t in trades:
        # Trade Section Header
        ws_boq.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        t_cell = ws_boq.cell(row=row_idx, column=1, value=f"{t.get('trade_code')} - {t.get('trade_name').upper()}")
        t_cell.font = white_bold_font
        t_cell.fill = trade_header_fill
        row_idx += 1

        for item in t.get("items", []):
            ws_boq.cell(row=row_idx, column=1, value=item.get("item_code")).font = bold_font
            ws_boq.cell(row=row_idx, column=2, value=item.get("name")).font = regular_font
            ws_boq.cell(row=row_idx, column=3, value=item.get("unit")).font = regular_font
            
            qty_cell = ws_boq.cell(row=row_idx, column=4, value=item.get("quantity"))
            qty_cell.font = regular_font
            qty_cell.number_format = '#,##0.0' if isinstance(item.get("quantity"), float) else '#,##0'
            
            rate_cell = ws_boq.cell(row=row_idx, column=5, value=item.get("rate"))
            rate_cell.font = regular_font
            rate_cell.number_format = '#,##0'

            # Dynamic Formula for Total Amount: =Quantity * Rate
            amt_cell = ws_boq.cell(row=row_idx, column=6, value=f"=D{row_idx}*E{row_idx}")
            amt_cell.font = bold_font
            amt_cell.number_format = '#,##0'

            ws_boq.cell(row=row_idx, column=7, value=item.get("layman_note", "")).font = subtitle_font
            row_idx += 1

    # -------------------------------------------------------------------------
    # TAB 3: SHOPPING LIST (CONSOLIDATED MATERIALS)
    # -------------------------------------------------------------------------
    ws_shop = wb.create_sheet(title="Consolidated Shopping List")
    ws_shop.views.sheetView[0].showGridLines = True

    ws_shop.cell(row=1, column=1, value="CONSOLIDATED MATERIAL PURCHASING LIST").font = title_font
    ws_shop.cell(row=2, column=1, value="Total aggregate quantities required to buy from suppliers.").font = subtitle_font

    shop_headers = ["Key Building Material", "Total Quantity Required", "Purchasing Notes"]
    for c_idx, h in enumerate(shop_headers, 1):
        cell = ws_shop.cell(row=4, column=c_idx, value=h)
        cell.font = white_bold_font
        cell.fill = header_fill

    shopping_dict = boq_data.get("shopping_list_summary", {})
    r = 5
    for mat, qty in shopping_dict.items():
        ws_shop.cell(row=r, column=1, value=mat).font = bold_font
        q_cell = ws_shop.cell(row=r, column=2, value=qty)
        q_cell.font = bold_font
        q_cell.number_format = '#,##0'
        ws_shop.cell(row=r, column=3, value="Order from verified local hardware / factory.").font = regular_font
        r += 1

    # -------------------------------------------------------------------------
    # TAB 4: LAYPERSON GUIDE
    # -------------------------------------------------------------------------
    ws_guide = wb.create_sheet(title="Layperson Construction Tips")
    ws_guide.views.sheetView[0].showGridLines = True

    ws_guide.cell(row=1, column=1, value="FUNDI LAYPERSON CONSTRUCTION GUIDE").font = title_font
    
    tips = [
        ("1. Cement Storage", "Always store cement bags on wooden pallets off the ground and cover with waterproof tarpaulin. Moist air hardens cement."),
        ("2. Concrete Mix Ratios", "Standard structural concrete ratio is 1 bag cement : 2 wheelbarrows sand : 4 wheelbarrows ballast (1:2:4)."),
        ("3. Anti-Termite Treatment", "Ensure anti-termite chemical is sprayed thoroughly on raw soil BEFORE laying the polythene DPM sheet."),
        ("4. Rebar Stirrup Spacing", "Columns and ring beams require stirrups spaced 150mm - 200mm apart to prevent steel bending."),
        ("5. Fundi Labor Payments", "Pay fundis based on verified completed milestones (e.g. slab poured, walling complete) rather than daily advance loans.")
    ]
    r = 3
    for title, desc in tips:
        ws_guide.cell(row=r, column=1, value=title).font = bold_font
        ws_guide.cell(row=r+1, column=1, value=desc).font = regular_font
        r += 3

    # Adjust Column Widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if not val_str.startswith("="):
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    wb.save(output_filepath)
    print(f"✅ Excel BOQ spreadsheet successfully generated at: {output_filepath}")
    return output_filepath
